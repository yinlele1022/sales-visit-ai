"""Excel audit logger for the sales-visit AI pipeline.

Every processed record is appended to an .xlsx file as a single row, creating
a permanent, human-readable audit trail that can be reviewed or imported into BI tools.

Columns:
    时间戳, 输入原文, 上下文, 销售姓名, 客户, 联系人, 拜访时间, 议题,
    模型输出(JSON), 使用模型, 是否Fallback, 是否通过, 需人工复核,
    人工复核原因, 错误详情, 省察轮数, 平均置信度, 处理耗时(秒), 飞书通知状态,
    归因分类
"""

import json
import os
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

@dataclass
class AuditEntry:
    """Single audit row."""

    transcript: str
    context: dict[str, Any] = field(default_factory=dict)
    extracted: dict[str, Any] = field(default_factory=dict)
    provider_used: str = ""
    fallback_used: bool = False
    passed: bool = False
    human_review_required: bool = False
    human_reason: str = ""
    errors: list[dict] = field(default_factory=list)
    introspection_rounds: int = 0
    confidence: dict[str, float] = field(default_factory=dict)
    elapsed_seconds: float = 0.0
    feishu_status: str = ""

    @property
    def timestamp(self) -> str:
        from datetime import datetime, timezone, timedelta

        tz = timezone(timedelta(hours=8))
        return datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")

    @property
    def failure_category(self) -> str:
        """三层归因：定位失败根因。
        大脑=模型能力不足 | 工具=解析/调用失败 | 输入=文本质量差 | 规则=业务校验拦截 | 正常=通过"""
        if self.passed:
            return "正常"
        if self.fallback_used:
            return "大脑（模型能力）"
        if self.errors:
            for e in self.errors:
                et = e.get("error_type", "")
                if et in ("解析错误", "JSON解析", "Schema验证"):
                    return "工具（解析失败）"
                if et in ("逻辑",):
                    return "规则（业务校验）"
            return "大脑（模型能力）"
        if self.human_review_required:
            if "文本" in self.human_reason or "无法解析" in self.human_reason:
                return "输入（文本质量）"
            return "大脑（模型能力）"
        return "正常"


class ExcelAuditLogger:
    """Append audit rows to an Excel file."""

    COLUMNS = [
        ("时间戳", 22),
        ("输入原文", 50),
        ("上下文", 30),
        ("销售姓名", 12),
        ("客户", 15),
        ("联系人", 12),
        ("拜访时间", 18),
        ("议题", 25),
        ("模型输出(JSON)", 60),
        ("使用模型", 15),
        ("是否Fallback", 14),
        ("是否通过", 12),
        ("需人工复核", 14),
        ("人工复核原因", 40),
        ("错误详情", 40),
        ("省察轮数", 12),
        ("平均置信度", 14),
        ("处理耗时(秒)", 14),
        ("飞书通知状态", 20),
        ("归因分类", 18),
    ]

    def __init__(self, path: Optional[str | Path] = None):
        if path is None:
            path = os.getenv("AUDIT_LOG_PATH") or Path(__file__).parent / "audit_logs" / "sales_visit_ai_log.xlsx"
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._lock_file = self.path.parent / ".write.lock"
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        """Create Excel file with header if it doesn't exist."""
        if self.path.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "处理日志"
        for col_idx, (header, width) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        wb.save(self.path)

    def _wait_for_lock(self, timeout: float = 30.0) -> bool:
        """Naive file lock to avoid concurrent writes from multiple workers."""
        start = time.time()
        while self._lock_file.exists():
            if time.time() - start > timeout:
                return False
            time.sleep(0.05)
        try:
            self._lock_file.write_text("locked", encoding="utf-8")
            return True
        except Exception:
            return False

    def _release_lock(self) -> None:
        try:
            if self._lock_file.exists():
                self._lock_file.unlink()
        except Exception:
            pass

    def _avg_confidence(self, confidence: dict[str, float]) -> float:
        if not confidence:
            return 0.0
        values = [v for v in confidence.values() if isinstance(v, (int, float))]
        return round(sum(values) / len(values), 3) if values else 0.0

    def _short_context(self, context: dict[str, Any]) -> str:
        parts = []
        if context.get("sales_name"):
            parts.append(f"销售:{context['sales_name']}")
        if context.get("reference_date"):
            parts.append(f"基准:{context['reference_date']}")
        if context.get("visit_time"):
            parts.append(f"拜访:{context['visit_time']}")
        return " | ".join(parts) or "（无）"

    def _short_errors(self, errors: list[dict]) -> str:
        if not errors:
            return ""
        parts = []
        for e in errors[:3]:
            t = e.get("error_type", "")
            f = e.get("field", "")
            r = e.get("reason", "")
            parts.append(f"{t}|{f}|{r}")
        return "; ".join(parts)

    def log(self, entry: AuditEntry) -> None:
        """Append a single audit row to the Excel file."""
        if not self._wait_for_lock():
            raise RuntimeError("Could not acquire write lock for audit log")
        try:
            wb = load_workbook(self.path)
            ws = wb.active
            row = ws.max_row + 1

            extracted_json = json.dumps(entry.extracted, ensure_ascii=False)
            if len(extracted_json) > 3000:
                extracted_json = extracted_json[:3000] + "...(truncated)"

            transcript_short = entry.transcript[:500] + ("..." if len(entry.transcript) > 500 else "")
            human_reason_short = entry.human_reason[:200]

            values = [
                entry.timestamp,
                transcript_short,
                self._short_context(entry.context),
                entry.extracted.get("sales_name", "未知") or "",
                entry.extracted.get("customer", "未知") or "",
                entry.extracted.get("contact_person", "未知") or "",
                entry.extracted.get("visit_time", "未知") or "",
                entry.extracted.get("topic", "未知") or "",
                extracted_json,
                entry.provider_used,
                "是" if entry.fallback_used else "否",
                "是" if entry.passed else "否",
                "是" if entry.human_review_required else "否",
                human_reason_short,
                self._short_errors(entry.errors),
                entry.introspection_rounds,
                self._avg_confidence(entry.confidence),
                round(entry.elapsed_seconds, 3),
                entry.feishu_status,
                entry.failure_category,
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx == 1:  # timestamp
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

            # Freeze header row
            ws.freeze_panes = "A2"
            wb.save(self.path)
        finally:
            self._release_lock()

    def log_batch(self, entries: list[AuditEntry]) -> None:
        """Append multiple rows efficiently."""
        if not self._wait_for_lock():
            raise RuntimeError("Could not acquire write lock for audit log")
        try:
            wb = load_workbook(self.path)
            ws = wb.active
            for entry in entries:
                row = ws.max_row + 1
                extracted_json = json.dumps(entry.extracted, ensure_ascii=False)
                if len(extracted_json) > 3000:
                    extracted_json = extracted_json[:3000] + "...(truncated)"
                transcript_short = entry.transcript[:500] + ("..." if len(entry.transcript) > 500 else "")
                human_reason_short = entry.human_reason[:200]
                values = [
                    entry.timestamp,
                    transcript_short,
                    self._short_context(entry.context),
                    entry.extracted.get("sales_name", "未知") or "",
                    entry.extracted.get("customer", "未知") or "",
                    entry.extracted.get("contact_person", "未知") or "",
                    entry.extracted.get("visit_time", "未知") or "",
                    entry.extracted.get("topic", "未知") or "",
                    extracted_json,
                    entry.provider_used,
                    "是" if entry.fallback_used else "否",
                    "是" if entry.passed else "否",
                    "是" if entry.human_review_required else "否",
                    human_reason_short,
                    self._short_errors(entry.errors),
                    entry.introspection_rounds,
                    self._avg_confidence(entry.confidence),
                    round(entry.elapsed_seconds, 3),
                    entry.feishu_status,
                    entry.failure_category,
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if col_idx == 1:
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"
            ws.freeze_panes = "A2"
            wb.save(self.path)
        finally:
            self._release_lock()

    def read_summary(self) -> dict[str, Any]:
        """Return basic statistics from the log."""
        if not self.path.exists():
            return {"total": 0, "passed": 0, "human_review": 0, "failed": 0}
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        total = max(0, ws.max_row - 1)
        passed = 0
        human = 0
        failed = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 13:
                continue
            if row[11] == "是":
                passed += 1
            if row[12] == "是":
                human += 1
            if row[11] == "否" and row[12] == "否":
                failed += 1
        return {"total": total, "passed": passed, "human_review": human, "failed": failed}


if __name__ == "__main__":
    # Quick self-test
    logger = ExcelAuditLogger()
    entry = AuditEntry(
        transcript="今天拜访了比亚迪张经理，聊了下高速 NOA。",
        context={"sales_name": "陈敏", "reference_date": "2026-06-22"},
        extracted={"sales_name": "陈敏", "customer": "比亚迪", "contact_person": "张经理", "topic": "高速 NOA"},
        provider_used="deepseek",
        fallback_used=False,
        passed=True,
        human_review_required=False,
        elapsed_seconds=1.23,
        feishu_status="无需通知",
    )
    logger.log(entry)
    print("Logged to:", logger.path)
    print("Summary:", logger.read_summary())
